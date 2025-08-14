from flask import Flask, request, jsonify
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

EXCEL_FILE = "submissions.xlsx"

# Create file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Message"])  # Column headers
    wb.save(EXCEL_FILE)

@app.route("/submit", methods=["POST"])
def submit_form():
    data = request.get_json()

    if not data or not data.get("name") or not data.get("email"):
        return jsonify({"error": "Name and Email are required"}), 400

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([data["name"], data["email"], data.get("message", "")])
    wb.save(EXCEL_FILE)

    return jsonify({"success": True, "message": "Form submitted!"}), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
